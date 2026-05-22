import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os
from pathlib import Path

DB_URL = os.environ['DATABASE_URL']
BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / 'outputs'
LEGACY_OUTPUT_DIR = BASE_DIR / 'scripts' / 'outputs'
os.makedirs(OUTPUT_DIR, exist_ok=True)

print('Pulling data from PostgreSQL...')
engine = create_engine(DB_URL)

with engine.connect() as conn:
    df       = pd.read_sql(text('SELECT * FROM lpl_ab_analysis'), conn)
    stats_df = pd.read_sql(text('SELECT * FROM lpl_stats_summary'), conn)

print(f'  {len(df)} game rows loaded')
print()

ab_stats_path = OUTPUT_DIR / 'lpl_ab_stats_results.csv'
if not ab_stats_path.exists():
    legacy_ab_stats_path = LEGACY_OUTPUT_DIR / 'lpl_ab_stats_results.csv'
    if legacy_ab_stats_path.exists():
        ab_stats_path = legacy_ab_stats_path

ab_df = pd.read_csv(ab_stats_path)

scenario_df = df.groupby(['side', 'scenario'])['result'].agg(
    win_rate='mean',
    games='count'
).round(4).reset_index()
scenario_df['win_rate_pct'] = (scenario_df['win_rate'] * 100).round(1)

pivot = scenario_df.pivot_table(
    index='scenario', columns='side', values='win_rate_pct'
).reset_index()

patch_df = df.groupby(['patch', 'side', 'got_fd'])['result'].mean().round(4).reset_index()
patch_df.columns = ['patch', 'side', 'got_fd', 'win_rate']
patch_df['win_rate_pct'] = (patch_df['win_rate'] * 100).round(1)
patch_df['group'] = patch_df.apply(
    lambda r: f"{r['side']} - {'Got FD' if r['got_fd']==1 else 'No FD'}", axis=1
)

print('Building Excel workbook...')
wb = Workbook()

BLUE_DARK   = '1A3C6B'
BLUE_MID    = '2E75B6'
BLUE_LIGHT  = 'BDD7EE'
RED_DARK    = '7B1515'
RED_MID     = 'C00000'
RED_LIGHT   = 'FFCCCC'
GREEN_LIGHT = 'E2EFDA'
GREY_LIGHT  = 'F2F2F2'
WHITE       = 'FFFFFF'
BLACK       = '000000'

def make_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def make_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(cell, bg=BLUE_DARK, fg=WHITE, bold=True, size=11):
    cell.fill      = make_fill(bg)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = make_border()

def data_style(cell, bg=WHITE, bold=False, align='center'):
    cell.fill      = make_fill(bg)
    cell.font      = Font(bold=bold, color=BLACK, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = make_border()

ws1 = wb.active
ws1.title = 'Executive Summary'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 22

ws1.merge_cells('A1:E1')
ws1['A1'] = 'LPL 2026 — First Dragon vs Void Grubs | A/B Test Results by Side'
ws1['A1'].font      = Font(bold=True, size=14, color=WHITE)
ws1['A1'].fill      = make_fill(BLUE_DARK)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:E2')
ws1['A2'] = 'Research Question: Does securing first dragon or void grubs impact win rate differently for Blue vs Red side in the LPL?'
ws1['A2'].font      = Font(italic=True, size=10, color=BLUE_DARK)
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 20

ws1.row_dimensions[3].height = 10

ws1.merge_cells('A4:E4')
ws1['A4'] = 'A/B TEST RESULTS'
header_style(ws1['A4'], bg=BLUE_MID)
ws1.row_dimensions[4].height = 22

headers = ['Side', 'Objective', 'Treatment Win Rate', 'Control Win Rate', 'Delta']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=5, column=col, value=h)
    header_style(cell, bg=BLUE_LIGHT, fg=BLACK)
ws1.row_dimensions[5].height = 18

ab_display = ab_df[['side','objective','treatment_wr','control_wr','delta','significant','p_value']].copy()
row = 6
for _, r in ab_display.iterrows():
    bg = GREY_LIGHT if row % 2 == 0 else WHITE
    ws1.cell(row=row, column=1, value=r['side'])
    ws1.cell(row=row, column=2, value=r['objective'])
    ws1.cell(row=row, column=3, value=f"{r['treatment_wr']*100:.1f}%")
    ws1.cell(row=row, column=4, value=f"{r['control_wr']*100:.1f}%")

    delta_cell = ws1.cell(row=row, column=5, value=f"+{r['delta']*100:.1f}%")
    sig_bg = GREEN_LIGHT if r['significant'] == 'YES' else RED_LIGHT
    for col in range(1, 6):
        data_style(ws1.cell(row=row, column=col), bg=bg)
    delta_cell.fill = make_fill(sig_bg)
    delta_cell.font = Font(bold=True, size=10,
                           color='375623' if r['significant']=='YES' else RED_DARK)
    ws1.row_dimensions[row].height = 18
    row += 1

ws1.row_dimensions[row].height = 10
row += 1
ws1.merge_cells(f'A{row}:E{row}')
ws1[f'A{row}'] = '✓ Green delta = statistically significant (p < 0.05)    ✗ Red delta = not significant'
ws1[f'A{row}'].font      = Font(italic=True, size=9, color='595959')
ws1[f'A{row}'].alignment = Alignment(horizontal='left')
row += 2

ws1.merge_cells(f'A{row}:E{row}')
ws1[f'A{row}'] = 'KEY FINDINGS'
header_style(ws1[f'A{row}'], bg=BLUE_MID)
ws1.row_dimensions[row].height = 22
row += 1

findings = [
    ('First Dragon', 'Statistically significant on BOTH sides (p < 0.001). +22.3pp win rate delta regardless of side.'),
    ('Void Grubs',   'NOT statistically significant on either side (p = 0.475). Grubs do not predict winning in the LPL.'),
    ('Blue Side',    'Gets first dragon only 38% of the time — but when they do, they win at 66% (highest scenario).'),
    ('Red Side',     'Gets first dragon 62% of the time — structural map advantage. Impact identical to Blue side.'),
    ('Conclusion',   'First dragon is ~2.7x more correlated with winning than grubs. Prioritise dragon over grubs in LPL draft.'),
]
for label, text in findings:
    ws1.merge_cells(f'B{row}:E{row}')
    label_cell = ws1.cell(row=row, column=1, value=label)
    text_cell  = ws1.cell(row=row, column=2, value=text)
    label_cell.font      = Font(bold=True, size=10, color=BLUE_DARK)
    label_cell.alignment = Alignment(vertical='center')
    label_cell.border    = make_border()
    text_cell.font       = Font(size=10)
    text_cell.alignment  = Alignment(wrap_text=True, vertical='center')
    text_cell.border     = make_border()
    ws1.row_dimensions[row].height = 28
    row += 1

ws2 = wb.create_sheet('AB Stats')
ws2.sheet_view.showGridLines = False

col_widths = [10, 16, 16, 16, 14, 14, 10, 10, 10, 12, 12]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:K1')
ws2['A1'] = 'Full A/B Test Statistics — LPL 2026'
ws2['A1'].font      = Font(bold=True, size=13, color=WHITE)
ws2['A1'].fill      = make_fill(BLUE_DARK)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

ab_headers = [
    'Side','Objective','Treatment Label','Control Label',
    'Treatment N','Control N','Treatment WR','Control WR',
    'Delta','P-Value','Significant'
]
for col, h in enumerate(ab_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    header_style(cell, bg=BLUE_LIGHT, fg=BLACK, size=10)
ws2.row_dimensions[2].height = 18

for r_idx, r in ab_df.iterrows():
    bg = GREY_LIGHT if r_idx % 2 == 0 else WHITE
    row_data = [
        r['side'], r['objective'], r['treatment_label'], r['control_label'],
        r['treatment_n'], r['control_n'],
        f"{r['treatment_wr']*100:.1f}%", f"{r['control_wr']*100:.1f}%",
        f"+{r['delta']*100:.1f}%", r['p_value'],
        r['significant']
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r_idx+3, column=col, value=val)
        data_style(cell, bg=bg)
        if col == 11:
            cell.fill = make_fill(GREEN_LIGHT if val == 'YES' else RED_LIGHT)
            cell.font = Font(bold=True, size=10,
                             color='375623' if val=='YES' else RED_DARK)
    ws2.row_dimensions[r_idx+3].height = 18

ws3 = wb.create_sheet('Scenario Breakdown')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18

ws3.merge_cells('A1:D1')
ws3['A1'] = 'Win Rate by Scenario and Side — LPL 2026'
ws3['A1'].font      = Font(bold=True, size=13, color=WHITE)
ws3['A1'].fill      = make_fill(BLUE_DARK)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

pivot_headers = ['Scenario', 'Blue Win Rate', 'Red Win Rate', 'Games (each side)']
for col, h in enumerate(pivot_headers, 1):
    cell = ws3.cell(row=2, column=col, value=h)
    header_style(cell, bg=BLUE_LIGHT, fg=BLACK, size=10)
ws3.row_dimensions[2].height = 18

scenarios = ['FD + More Grubs', 'FD Only', 'More Grubs Only', 'Neither']
scenario_colors = [GREEN_LIGHT, BLUE_LIGHT, GREY_LIGHT, RED_LIGHT]

for r_idx, (sc, bg) in enumerate(zip(scenarios, scenario_colors)):
    row_num = r_idx + 3
    blue_row = scenario_df[(scenario_df['scenario']==sc) & (scenario_df['side']=='Blue')]
    red_row  = scenario_df[(scenario_df['scenario']==sc) & (scenario_df['side']=='Red')]
    blue_wr  = f"{blue_row['win_rate_pct'].values[0]:.1f}%" if len(blue_row) else 'N/A'
    red_wr   = f"{red_row['win_rate_pct'].values[0]:.1f}%"  if len(red_row)  else 'N/A'
    games    = int(blue_row['games'].values[0]) if len(blue_row) else 0

    for col, val in enumerate([sc, blue_wr, red_wr, games], 1):
        cell = ws3.cell(row=row_num, column=col, value=val)
        data_style(cell, bg=bg)
        if col == 1:
            cell.font = Font(bold=True, size=10)
    ws3.row_dimensions[row_num].height = 20

ws4 = wb.create_sheet('Patch Trends')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 24
ws4.column_dimensions['C'].width = 16

ws4.merge_cells('A1:C1')
ws4['A1'] = 'First Dragon Win Rate by Patch and Side — LPL 2026'
ws4['A1'].font      = Font(bold=True, size=13, color=WHITE)
ws4['A1'].fill      = make_fill(BLUE_DARK)
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

for col, h in enumerate(['Patch', 'Group', 'Win Rate'], 1):
    cell = ws4.cell(row=2, column=col, value=h)
    header_style(cell, bg=BLUE_LIGHT, fg=BLACK, size=10)
ws4.row_dimensions[2].height = 18

patch_display = patch_df[['patch','group','win_rate_pct']].sort_values(['patch','group'])
for r_idx, (_, r) in enumerate(patch_display.iterrows()):
    row_num = r_idx + 3
    bg = GREY_LIGHT if r_idx % 2 == 0 else WHITE
    for col, val in enumerate([r['patch'], r['group'], f"{r['win_rate_pct']:.1f}%"], 1):
        cell = ws4.cell(row=row_num, column=col, value=val)
        data_style(cell, bg=bg)
    ws4.row_dimensions[row_num].height = 18

output_path = OUTPUT_DIR / 'lpl_executive_summary.xlsx'
wb.save(output_path)
print(f'Excel workbook saved to {output_path}')
print()
print('Tabs created:')
print('  1. Executive Summary  — key findings + color-coded A/B results')
print('  2. A/B Stats          — full statistical output table')
print('  3. Scenario Breakdown — win rate by all 4 scenarios x side')
print('  4. Patch Trends       — FD win rate across patches by side')
print()
print('All done. Load outputs/ folder into Power BI next.')
